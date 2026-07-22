"""Binance 资金费率狙击: 裸合约单向持仓，收取极端费率后立即平仓。"""
from __future__ import annotations
import asyncio
import csv
import gc
import math
import shutil
import time
import uuid
from pathlib import Path
from typing import Any, TYPE_CHECKING

import pandas as pd

from ..constants import *
from .. import constants as _c
from ..models import LegResult, _safe_float

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, numbers
    from openpyxl.utils import get_column_letter
    _XLSX_AVAILABLE = True
except ImportError:
    _XLSX_AVAILABLE = False

if TYPE_CHECKING:
    from ..bot import FundingArbitrageBot


class BnSnipeStrategy:
    """Binance 资金费率狙击策略。扫描全市场 USDT 永续，选取 abs(费率) 最高、
    覆盖 2x 手续费且通过流动性筛选的币种，在结算前瞬间开裸单，
    结算时间到立即平仓。"""

    def __init__(self, bot: FundingArbitrageBot) -> None:
        self.bot: FundingArbitrageBot = bot
        self.is_open: bool = False
        self.symbol: str = ""
        self.amount: float = 0.0
        self.direction: str = ""  # "short" 或 "long"
        self.next_funding_time_ms: float = 0.0
        self.entry_price: float = 0.0
        self.funding_rate: float = 0.0
        self._next_snipe_settle_ms: float = 0.0
        self._snipe_loop_guard: bool = False
        self.is_sniping: bool = False       # 并发穿透锁，防外部定时器重入
        self.current_delay_ms: float = _c.CROSS_SNIPER_CLOSE_DELAY_MS
        self.ws_funding_arrived_event = asyncio.Event()  # WS 资费到账信号
        self.last_funding_tx_ms: int = 0               # 币安官方结算毫秒戳，由 ws.py 写入
        self.last_funding_event_ms: int = 0            # 币安 WS 推送事件毫秒戳，由 ws.py 写入
        self.last_funding_recv_ms: int = 0             # 本地收到 WS 推送毫秒戳，由 ws.py 写入
        self.last_funding_amount_usdt: float = 0.0     # 资费到账金额(WS bc)，由 ws.py 写入，最准
        self.net_rate: float = 0.0
        self.futures_fee: float = 0.0
        self.bnb_price: float = 0.0                    # BNBUSDT 标记价，扫描时刷新，用于 BNB 手续费换算 USDT
        self._last_trigger_type: str = ""              # 本轮平仓触发方式: WS_TRIGGER / TIMEOUT_兜底
        self._last_e2e_ms: int = 0                     # 本轮穿透全链路耗时(币安记账→平仓发单)
        self.last_realized_pnl_usdt: float = 0.0       # 价格盈亏(WS rp 累加，交易所结算)，由 ws.py 写入
        self.last_commission: float = 0.0              # 手续费原生金额累加(WS n)，由 ws.py 写入
        self.last_commission_asset: str = ""           # 手续费币种(WS N)，由 ws.py 写入
        self.last_fill_count: int = 0                  # 本轮成交笔数(WS TRADE 计数)，由 ws.py 写入
        self._clean_symbol: str = ""                   # 当前狙击标的裸符号(GWEIUSDT)，供 ws.py 匹配成交
        self.tp_filled: bool = False                   # 本轮止盈单是否已成交，由 ws.py 写入（供账本打标签）
        self.last_open_avg_price: float = 0.0          # 开仓腿真实成交均价(WS ap)，由 ws.py 写入，供止盈按真实价算触发
        self._tp_task: asyncio.Task | None = None      # 止盈挂单的 fire-and-forget 任务句柄（持引用防 GC）
        self._tp_client_id: str = ""                # 本轮止盈单 clientAlgoId，供平仓后主动撤单
        self._close_fill_arrived: bool = False      # 关仓腿 ORDER_TRADE_UPDATE 已到达，由 ws.py 写入（供记账轮询）
        self._interval_hours: int = 0               # 本轮标的资费周期(1/4/8h)，扫描时写入，供账本
        self._batch_size: int = 0                   # 本轮结算时刻全市场同批结算的币数（记账排队长度），供账本
        self._subset_rank: int = -1                 # 本轮标的在同批(同结算时刻)字母序中的位次(0起)，供账本
        self._settle_hour_utc: int = -1             # 本轮结算节点 UTC 整点(0-23)，供账本
        self._funding_intervals: dict[str, int] = {}  # 资费周期缓存(原生符号→小时)，1h 刷新
        self._funding_intervals_at: float = 0.0
        # ── 全链路时间戳（绝对值，CSV 审计用） ──
        self._open_send_time: float = 0.0           # 发出开仓订单 (time.time())
        self._open_fill_time_ms: int = 0            # 开仓成交时刻 (ORDER_TRADE_UPDATE T)，由 ws.py 写入
        self._open_ack_time: float = 0.0            # 收到开仓确认 (time.time())
        self._tp_send_time: float = 0.0             # 发出止盈单 (time.time())
        self._close_send_time: float = 0.0          # 发出平仓订单 (time.time())
        self._close_fill_time_ms: int = 0           # 平仓成交时刻 (ORDER_TRADE_UPDATE T)，由 ws.py 写入
        self._close_ack_time: float = 0.0           # 收到平仓确认 (time.time())

    # ── 扫描 ───────────────────────────────────────────────

    async def scan(self, position_usdt: float = 100.0) -> pd.DataFrame:
        """扫描全市场 USDT 永续，按 abs(资金费率) 降序排列。"""
        futures_tickers, funding_rates, fee_pair = await asyncio.gather(
            self.bot._safe_request(
                "futures.fetch_tickers",
                lambda: self.bot.futures.fetch_tickers(),
                default={},
            ),
            self.bot._safe_request(
                "futures.fetch_funding_rates",
                lambda: self.bot.futures.fetch_funding_rates(),
                default={},
            ),
            self.bot.fetch_taker_fees(),
        )
        _spot_fees, futures_taker_fees = fee_pair
        default_futures_fee = futures_taker_fees.get("__default__", DEFAULT_FUTURES_TAKER_FEE)

        # 资费周期：exchangeInfo 不含 fundingIntervalHours，需查 fundingInfo（公开接口，不用签名）。
        # 缓存1h：狙击扫描在 T-10s、总预算12s，这里若慢请求会挤占开仓窗口，短超时+缓存兜底。
        if not self._funding_intervals or time.time() - self._funding_intervals_at > 3600:
            try:
                fi_data = await asyncio.wait_for(
                    self.bot._binance_public_get("/fapi/v1/fundingInfo"), timeout=3,
                )
                intervals: dict[str, int] = {}
                for fi in fi_data:
                    sym = fi.get("symbol", "")
                    interval = fi.get("fundingIntervalHours")
                    if sym and interval is not None:
                        intervals[sym] = int(interval)
                if intervals:
                    self._funding_intervals = intervals
                    self._funding_intervals_at = time.time()
            except Exception:
                pass  # 失败沿用旧缓存/默认8h，不阻塞扫描
        funding_intervals = self._funding_intervals

        futures_market_index = self.bot._build_futures_market_index()

        # 字母排位：该 base 在全量 USDT 永续字母序中的百分位（0=最靠前，→1=最靠后）。
        # 用于动态时点——字母越靠后，币安资费记账越晚（实测 Pearson≈0.77）。
        all_bases = sorted(futures_market_index.keys())
        n_bases = max(1, len(all_bases))
        base_rank_pct = {b: i / n_bases for i, b in enumerate(all_bases)}

        # 批：全市场每个结算时刻(next_funding_time)有哪些币一起结算，按字母序排。
        # 币安按批、且批内按字母序处理记账 → 批越大排队越长、批内越靠后记账越晚。
        # batch_size=该批币数；subset_rank=本币在该批字母序中的位次(0起)——比全市场 rank_pct 更准。
        nft_batch_members: dict[int, list[str]] = {}
        for _base, _m in futures_market_index.items():
            _nft = self.bot._extract_next_funding_time(funding_rates.get(_m["symbol"], {}))
            if _nft:
                nft_batch_members.setdefault(_nft, []).append(_base)
        nft_batch_pos: dict[int, dict[str, int]] = {}
        for _nft, _lst in nft_batch_members.items():
            _lst.sort()
            nft_batch_pos[_nft] = {b: i for i, b in enumerate(_lst)}

        # 记录 BNBUSDT 标记价，用于把 BNB 抵扣的手续费换算成 USDT
        bnb_market = futures_market_index.get("BNB")
        if bnb_market:
            bnb_last = _safe_float(futures_tickers.get(bnb_market["symbol"], {}).get("last"))
            if bnb_last > 0:
                self.bnb_price = bnb_last

        rows: list[dict[str, Any]] = []
        for base, market in futures_market_index.items():
            symbol = market["symbol"]
            ticker = futures_tickers.get(symbol, {})
            quote_volume = _safe_float(ticker.get("quoteVolume"))
            if not self.bot._futures_passes_volume(quote_volume):
                continue

            funding_item = funding_rates.get(symbol, {})
            rate, _is_predicted = self.bot._extract_predicted_funding_rate(funding_item)
            if rate is None:
                continue

            next_ft = self.bot._extract_next_funding_time(funding_item)
            abs_rate = abs(rate)
            futures_fee = futures_taker_fees.get(symbol, default_futures_fee)
            net_rate = abs_rate - 2 * futures_fee
            direction = "short" if rate > 0 else "long"

            passes = (
                net_rate > 0
                and abs_rate >= _c.BN_SNIPE_MIN_ABS_RATE
            )

            remain_s = (next_ft - time.time() * 1000) / 1000 if (next_ft is not None and next_ft > 0) else 0
            settle_str = time.strftime("%H:%M", time.localtime(next_ft / 1000)) if (next_ft is not None and next_ft > 0) else "?"

            # fundingInfo 返回原生符号(BTCUSDT)，用 market["id"] 查，别用统一格式 symbol
            interval_hours = funding_intervals.get(market.get("id", ""), DEFAULT_FUNDING_INTERVAL_HOURS)

            rows.append({
                "symbol": symbol,
                "base": base,
                "rank_pct": base_rank_pct.get(base, 0.0),
                "futures_price": _safe_float(ticker.get("last")),
                "funding_rate": rate,
                "abs_rate": abs_rate,
                "net_rate": net_rate,
                "direction": direction,
                "next_funding_time_ms": next_ft,
                "futures_taker_fee": futures_fee,
                "passes": passes,
                "quote_volume": quote_volume,
                "remain_min": remain_s / 60,
                "settle_local": settle_str,
                "interval_hours": interval_hours,
                "batch_size": len(nft_batch_members.get(next_ft, [])),
                "subset_rank": nft_batch_pos.get(next_ft, {}).get(base, -1),
                "contract_type": market.get("info", {}).get("contractType", ""),
            })

        if not rows:
            return pd.DataFrame()

        df = pd.DataFrame(rows)
        df.sort_values(["next_funding_time_ms", "abs_rate"], ascending=[True, False], inplace=True)
        df.reset_index(drop=True, inplace=True)
        return df

    # ── 开仓 ───────────────────────────────────────────────

    async def open_position_fast(self, args: dict) -> bool:
        """极速开仓：接收 T-10s 预解包的原生 dict，零 Pandas 开销直发 WS。"""
        symbol = args["symbol"]
        direction = args["direction"]
        amount = args["amount"]
        rate = args["rate"]
        nft = args["nft"]
        price = args["price"]

        # 重置本轮账本采集字段，防上一轮残留污染延迟/触发统计
        self.last_funding_tx_ms = 0
        self.last_funding_event_ms = 0
        self.last_funding_recv_ms = 0
        self.last_funding_amount_usdt = 0.0
        self._last_trigger_type = ""
        self._last_e2e_ms = 0
        # 重置成交回报采集 + 设定匹配符号（须在开仓发单前，确保开仓成交也被 ws.py 捕获）
        self.last_realized_pnl_usdt = 0.0
        self.last_commission = 0.0
        self.last_commission_asset = ""
        self.last_fill_count = 0
        self.tp_filled = False
        self.last_open_avg_price = 0.0
        self._tp_client_id = ""
        self._tp_task = None
        self._close_fill_arrived = False
        self._open_send_time = 0.0
        self._open_fill_time_ms = 0
        self._open_ack_time = 0.0
        self._tp_send_time = 0.0
        self._close_send_time = 0.0
        self._close_fill_time_ms = 0
        self._close_ack_time = 0.0
        self._clean_symbol = self.bot._clean_futures_symbol(symbol)

        send_ts = time.time()
        self._open_send_time = send_ts
        t0 = time.perf_counter()
        # 开仓作为后台任务跑（其内含 ~50ms 持仓校验 sleep）；不等它整条返回——
        # 一旦 WS 捕获到开仓真实成交均价(ap，约 :00.01 到)就立刻按真实价挂止盈，
        # 把持仓校验的 ~50ms 与挂止盈并行掉：止盈约 :00.03 就能挂上（≈旧并发的速度，但用真实价算触发）。
        if direction == "short":
            open_coro = self.bot._cross_open_short_leg("binance", symbol, amount, skip_position_check=True)
            pos_side = "SHORT"
        else:
            open_coro = self.bot._cross_open_long_leg("binance", symbol, amount, skip_position_check=True)
            pos_side = "LONG"
        open_task = asyncio.create_task(open_coro)

        # 等真实成交价到手（WS ap），最多 ~150ms；到手即按真实价挂止盈
        tp_client_id = ""
        if _c.BN_SNIPE_TAKE_PROFIT_ENABLED and rate:
            for _ in range(150):
                if self.last_open_avg_price > 0 or open_task.done():
                    break
                await asyncio.sleep(0.001)
            entry_now = self.last_open_avg_price
            if entry_now > 0:   # ap>0 即证明开仓单已成交、仓位已存在 → 挂止盈安全
                # 止盈触发距 = abs(费率)×倍数 → 命中即锁≥毛资费
                tp_dist = abs(rate) * _c.BN_SNIPE_TP_RATE_MULT
                if direction == "long":
                    tp_side, raw_stop = "sell", entry_now * (1 + tp_dist)
                else:
                    tp_side, raw_stop = "buy", entry_now * (1 - tp_dist)
                try:
                    # price_to_precision 返回字符串（已按 tick 取整，且避免低价币科学计数法）
                    stop_price = self.bot.futures.price_to_precision(symbol, raw_stop)
                except Exception:
                    stop_price = f"{raw_stop:.8f}"
                cid = "TPSNIPE" + uuid.uuid4().hex[:8]
                # fire-and-forget：止盈是尽力而为的旁挂，绝不能让它的回执卡在半路而
                # 拖住关键的 开仓→等资费→平仓 路径（await 回执最坏会堵到 5s 超时）。
                # 单子已在 send_json 那刻发出、由该协程自行 log 成败；主流程不等它。
                self._tp_send_time = time.time()
                self._tp_task = asyncio.create_task(
                    self.bot._bn_trade_ws_stop_order(
                        symbol, tp_side, stop_price, pos_side, cid))
                tp_client_id = cid
                self._tp_client_id = cid

        result = await open_task
        self._open_ack_time = time.time()
        rtt_ms = (time.perf_counter() - t0) * 1000
        send_str = time.strftime("%H:%M:%S", time.localtime(send_ts))
        send_str += f".{int((send_ts % 1) * 1000):03d}"

        if not result.ok:
            logger.warning("[狙击] 开仓失败 %s: %s | 发送=%s | RTT=%.1fms",
                           symbol, result.error, send_str, rtt_ms)
            # 开仓失败但止盈已挂上 → 尽力撤掉，避免残留僵尸挂单
            if tp_client_id:
                await self.bot._bn_cancel_order(symbol, tp_client_id)
            return False

        # 真实开仓均价（记账/状态用）：优先 WS ap，退回下单响应 avgPrice，再退回扫描价
        entry_fill = self.last_open_avg_price
        if entry_fill <= 0 and isinstance(result.order, dict):
            try:
                entry_fill = float((result.order.get("info") or {}).get("avgPrice", 0) or 0)
            except (TypeError, ValueError, AttributeError):
                entry_fill = 0.0
        if entry_fill <= 0:
            entry_fill = price

        self.is_open = True
        self.symbol = symbol
        self.amount = amount
        self.direction = direction
        self.next_funding_time_ms = nft
        self.entry_price = entry_fill
        self.funding_rate = rate
        self.net_rate = float(args.get("net_rate", 0))
        self.futures_fee = float(args.get("futures_fee", DEFAULT_FUTURES_TAKER_FEE))
        logger.info("[狙击] 开仓成功 %s: %s %.4f张 @ %.6f 费率=%.4f%% | 发送=%s | RTT=%.1fms",
                     symbol, direction, amount, entry_fill, rate * 100, send_str, rtt_ms)
        return True

    # ── 平仓 ───────────────────────────────────────────────

    async def close_position_fast(self) -> tuple[bool, float]:
        """极速平仓。先开枪后说话，重试间隔 2ms。返回 (成功, 纯净RTT_ms)。
        ReduceOnly 拒绝 = 首次已成交，视为成功。
        止盈已满额成交时跳过发单，直接收尾。"""
        if not self.is_open:
            return True, 0.0

        sym, amt, d = self.symbol, self.amount, self.direction
        send_ts = time.time()
        self._close_send_time = send_ts
        t0 = time.perf_counter()

        # 止盈已在机房侧完全成交(closePosition=true, 仓位归零) → 跳过市价平仓，
        # 不发任何冗余单，直接进入状态清理与账本落盘。cached result 保证下游不变。
        if self.tp_filled:
            logger.info("[狙击] 止盈已在机房侧完全成交，跳过市价平仓发单")
            result: LegResult = LegResult(True, "futures", sym,
                                          "BUY" if d == "short" else "SELL", amt)
            rtt_ms = 0.0
        else:
            for attempt in range(3):
                if d == "short":
                    result = await self.bot._cross_close_short_leg("binance", sym, amt)
                else:
                    result = await self.bot._cross_close_long_leg("binance", sym, amt)
                if result.ok:
                    break
                # ReduceOnly = 仓位已被首次平掉，视为成功
                if result.error and "-2022" in str(result.error):
                    logger.info("[狙击] 平仓重试命中 ReduceOnly，首次已成交，视为成功")
                    result = LegResult(True, result.market_type, result.symbol, result.side, result.amount)
                    break
                if attempt < 2:
                    logger.warning("[狙击] 平仓受阻！第 %d/3 次极速重试", attempt + 1)
                    await asyncio.sleep(0.002)
            rtt_ms = (time.perf_counter() - t0) * 1000

        send_str = time.strftime("%H:%M:%S", time.localtime(send_ts))
        send_str += f".{int((send_ts % 1) * 1000):03d}"

        if not result.ok:
            logger.critical("[狙击] 平仓失败！%s | 发送=%s", result.error, send_str)
            return False, rtt_ms

        self._close_ack_time = time.time()

        # 只有确认平仓成功后才清状态，防止幽灵裸仓
        # 记录前先保存值。此时 WS 已收到开仓真实均价 ap，用它覆写 _entry 得精确名义价值
        _sym, _amt, _entry, _dir = sym, amt, self.entry_price, d
        if self.last_open_avg_price > 0:
            _entry = self.last_open_avg_price
        _fee = self.futures_fee
        _nft = self.next_funding_time_ms
        _rate = self.funding_rate

        self.is_open = False
        self.symbol = ""
        self.amount = 0.0
        self.direction = ""
        self.next_funding_time_ms = 0.0
        self.entry_price = 0.0
        self.funding_rate = 0.0
        self.net_rate = 0.0
        self.futures_fee = 0.0

        # 主动撤止盈单兜底：仓位已平。先 await tp_task 确认止盈单确已落地（内部自带5s上限、不抛），
        # 再按 clientAlgoId 精撤——补上"平仓抢在止盈落地之前"的僵尸单缺口。
        # 与币安 closePosition 自动撤单叠加：正常路径下自动撤单已撤 → 这刀撤到 unknown order，被 _bn_cancel_order 吞掉，幂等无害。
        if self._tp_task is not None:
            try:
                await self._tp_task
            except Exception:
                pass
            if self._tp_client_id:
                await self.bot._bn_cancel_order(_sym, self._tp_client_id)
            self._tp_task = None

        # 平仓后算完整利润：价格盈亏 + 资费 - 手续费
        buy_avg = sell_avg = _entry
        total_commission = 0.0
        commission_asset = ""
        trade_count = 0
        # 先等资费真正到账再记账：大结算(00/08/16 UTC)币安推送可能晚 1-2s，
        # 若过早查账会把资费记成 0、平仓腿也可能还没落成交。仓位已平，等的是记账不影响交易。
        # 止盈已成交则不会有资费到账，跳过等待免白等 5s。
        if not self.tp_filled and not self.ws_funding_arrived_event.is_set():
            try:
                await asyncio.wait_for(self.ws_funding_arrived_event.wait(), timeout=5.0)
            except asyncio.TimeoutError:
                logger.warning("[狙击] 资费 5s 未到账，按 REST 兜底记账 %s", _sym)
        await asyncio.sleep(0.5)

        # 关仓 ORDER_TRADE_UPDATE 在大结算(00/08/16 UTC)可能延迟 >1s 推送
        # （币安 E 字段滞后 T 字段 1.2s+）。轮询等到 _close_fill_arrived 确保
        # rp 已完整捕获后再记账，超时 5s 走 REST 兜底。
        # tp_filled 同理：止盈成交也会设 _close_fill_arrived（R=true），轮询即可。
        if not self._close_fill_arrived:
            for _ in range(50):
                if self._close_fill_arrived:
                    break
                await asyncio.sleep(0.1)
            if not self._close_fill_arrived:
                logger.warning("[狙击] 关仓成交推送未到达，改用 REST 兜底记账 %s", _sym)

        # ── 价格盈亏 + 手续费：优先 WS 实时成交回报(rp/n)，免 REST userTrades 竞态 ──
        # 拥堵结算下 REST 可能只返回单腿成交，rp/手续费漏半(曾把 -0.36 价盈记成 -0.036)。
        # WS ORDER_TRADE_UPDATE 每笔成交都推 rp(交易所结算价盈)与 n(手续费)，最准。
        # 必须等关仓成交推送已到（_close_fill_arrived）才能用 WS 数据，否则 rp 只有开仓腿(0)
        # → 价格盈亏记成 0。
        if self.last_fill_count > 0 and self._close_fill_arrived:
            price_pnl = self.last_realized_pnl_usdt
            total_commission = self.last_commission
            commission_asset = self.last_commission_asset
            trade_count = self.last_fill_count
            sell_avg = _entry + (price_pnl / _amt if _amt else 0.0)  # 反推平仓均价，仅供日志展示
            logger.debug("[手续费] WS成交 %d 笔 | rp合计=%.6f | commission=%.8f %s",
                          trade_count, price_pnl, total_commission, commission_asset)
        else:
            # WS 未捕获成交(资费WS断线等) → REST userTrades 兜底重建均价
            price_pnl = 0.0
            try:
                clean = self.bot._clean_futures_symbol(_sym)
                now_ms = int(time.time() * 1000)
                trades = await self.bot._binance_request(
                    BINANCE_FUTURES_API, "/fapi/v1/userTrades",
                    {"symbol": clean, "startTime": now_ms - 15000, "endTime": now_ms, "limit": 100},
                )
                buys: list[float] = []
                sells: list[float] = []
                if isinstance(trades, list):
                    for t in trades:
                        qty = float(t.get("qty", 0) or 0)
                        if qty <= 0:
                            continue
                        price = float(t["price"])
                        side = t.get("side", "")
                        if side == "BUY":
                            buys.append(price)
                        elif side == "SELL":
                            sells.append(price)
                        comm = float(t.get("commission", 0) or 0)
                        total_commission += comm
                        if comm and not commission_asset:
                            commission_asset = t.get("commissionAsset", "")

                buy_avg = sum(buys) / len(buys) if buys else _entry
                sell_avg = sum(sells) / len(sells) if sells else _entry
                trade_count = len(buys) + len(sells)
                logger.debug("[手续费] REST兜底 userTrades 返回 %d 笔 | commission 合计=%.8f %s",
                              trade_count, total_commission, commission_asset)
            except Exception as e:
                logger.debug("[手续费] userTrades 查询失败: %s", e)
            price_pnl = _amt * (sell_avg - buy_avg)  # long买→卖, short卖→买，同公式

        # 资费金额优先用 WS 的 bc（最准，已确认到账，正负皆可）；WS 未到账才退回 REST
        if self.ws_funding_arrived_event.is_set():
            funding_income = self.last_funding_amount_usdt
        else:
            funding_income = await self.bot._query_bn_funding_amount(_sym, int(_nft))

        # ── 手续费换算 USDT（本账户用 BNB 抵扣手续费）──
        fee_native = total_commission
        fee_asset = commission_asset
        if fee_native <= 0:
            # API 未返回手续费 → 用费率估算（USDT 计价）
            fee_native = _amt * _entry * 2 * _fee
            fee_asset = "USDT_EST"
            logger.debug("[手续费] API 返回 0，用费率 %.4f%% 估算=%.6f", _fee * 100, fee_native)

        bnb_price = self.bnb_price
        if fee_asset == "BNB":
            if bnb_price > 0:
                fee_usdt = fee_native * bnb_price
            else:
                # 无 BNBUSDT 报价，退回费率估算，避免把 BNB 数量误当 USDT
                fee_usdt = _amt * _entry * 2 * _fee
                logger.warning("[手续费] BNB 抵扣但无 BNBUSDT 报价，退回费率估算=%.6f", fee_usdt)
        else:
            # USDT / USDC / FDUSD / 估算 → 已是 USDT 计价
            fee_usdt = fee_native

        # price_pnl 已由上方 WS(rp) 或 REST(均价) 分支算得
        notional = _amt * _entry
        profit = price_pnl + funding_income - fee_usdt
        net_rate = profit / notional if notional > 0 else 0.0

        logger.info("[狙击] %s | 开=%.6f 平=%.6f | 价格盈亏=%.4f | 资费=%.4f | 手续费=%.6f %s(≈%.4fU) | 净利=%.4f (%.4f%%)",
                     _sym, buy_avg, sell_avg, price_pnl, funding_income,
                     fee_native, fee_asset, fee_usdt, profit, net_rate * 100)
        if funding_income <= 0:
            logger.warning("[狙击] 资费未到账！%s", _sym)

        self.bot._record_trade(_sym, _dir, profit, net_rate,
                                amount=notional,
                                price_pnl=price_pnl,
                                funding_pnl=funding_income,
                                fee_total=fee_usdt)

        # ── 完整账本落盘（真实盈亏 + 三段延迟 + 成交质量）──
        if self.tp_filled:
            self._last_trigger_type = "TP_止盈"

        # 方向感知的开仓价/平仓价/滑点
        _entry_price = _entry
        if _dir == "long":
            _exit_price = _entry + price_pnl / _amt if _amt else _entry
        else:
            _exit_price = _entry - price_pnl / _amt if _amt else _entry
        _slippage_pct = price_pnl / notional * 100 if notional else 0.0

        self._record_ledger(
            symbol=_sym, direction=_dir, funding_rate=_rate, notional=notional,
            entry_price=_entry_price, exit_price=_exit_price, slippage_pct=_slippage_pct,
            funding_income=funding_income, price_pnl=price_pnl,
            fee_native=fee_native, fee_asset=fee_asset, bnb_price=bnb_price,
            fee_usdt=fee_usdt, net_pnl=profit, fill_count=trade_count,
            settle_fallback_ms=int(_nft),
        )

        logger.info("[狙击] 平仓成功 %s %s %.4f张 | 发送=%s | RTT=%.1fms",
                     _sym, _dir, _amt, send_str, rtt_ms)

        # ── 全链路时间戳审计 ──
        order = result.order or {}
        fill_ts = order.get("transactTime", 0)
        self._log_full_timeline(fill_ts, send_ts)

        return True, rtt_ms

    # ── 辅助 ───────────────────────────────────────────────

    def should_exit(self) -> bool:
        """结算时间+延时已过则退出。与自旋锁共用同一延时，防止顶部持仓管理提前抢跑。"""
        if not self.is_open:
            return False
        if self.next_funding_time_ms <= 0:
            return True
        return time.time() * 1000 >= (self.next_funding_time_ms + self.current_delay_ms)

    async def check_position(self) -> bool:
        """裸仓安全守卫：WS 缓存优先，抖动时走 REST 最终确认，绝不盲清状态。"""
        if not self.is_open:
            return False
        pos_side = "SHORT" if self.direction == "short" else "LONG"
        ok, _ = self.bot._ws_position_check("binance", self.symbol, pos_side, expect_zero=False, amount=self.amount)
        if not ok:
            # WS 没查到不慌：高负载结算期 ACCOUNT_UPDATE 推送经常延迟数百 ms，
            # 用 REST 做最终确认——裸仓策略绝不能把真实的 WS 延迟误判为"仓位不存在"。
            pos = await self.bot._cross_verify_binance_position(self.symbol, pos_side)
            if pos > 0:
                logger.info("[狙击] WS持仓延迟，经 REST 确认仓位仍在 %.4f", pos)
                return True
            logger.warning("[狙击] 经 REST 最终确认仓位已不存在 %s %s", self.symbol, pos_side)
            self.is_open = False
            return False
        return True

    # ── 性能黑匣子 ─────────────────────────────────────────

    def _log_full_timeline(self, fill_ts: int, send_ts: float) -> None:
        """全链路时间戳审计：币安结算 → WS推送 → 本地收到 → 平仓发出 → 订单成交。"""
        ts = {
            "A_结算": self.last_funding_tx_ms,
            "B_WS推送": self.last_funding_event_ms,
            "C_本地收到": self.last_funding_recv_ms,
            "D_平仓发出": int(send_ts * 1000),
            "E_订单成交": int(fill_ts),
        }
        # 只打有效时间戳
        valid = {k: v for k, v in ts.items() if v > 0}
        if len(valid) < 2:
            return
        sorted_ts = sorted(valid.values())
        base = sorted_ts[0]
        timeline = " → ".join(f"{k}={v - base:+d}ms" for k, v in ts.items() if v > 0)
        logger.info("[全链路] %s", timeline)

    _BJT = 8 * 3600  # UTC+8 偏移，强制北京时间

    @staticmethod
    def _fmt_ms(ms: int) -> str:
        """毫秒时间戳 → 北京时间字符串，0 返回空。"""
        if ms <= 0:
            return ""
        lt = time.gmtime(ms / 1000 + BnSnipeStrategy._BJT)
        return time.strftime("%Y-%m-%d %H:%M:%S", lt) + f".{ms % 1000:03d}"

    @staticmethod
    def _fmt_ts(ts: float) -> str:
        """time.time() 浮点秒 → 北京时间字符串，0 返回空。"""
        if ts <= 0:
            return ""
        lt = time.gmtime(ts + BnSnipeStrategy._BJT)
        return time.strftime("%Y-%m-%d %H:%M:%S", lt) + f".{int((ts % 1) * 1000):03d}"

    def _record_ledger(self, *, symbol: str, direction: str, funding_rate: float,
                       notional: float, entry_price: float, exit_price: float,
                       slippage_pct: float, funding_income: float, price_pnl: float,
                       fee_native: float, fee_asset: str, bnb_price: float,
                       fee_usdt: float, net_pnl: float, fill_count: int,
                       settle_fallback_ms: int) -> None:
        """完整账本：一单一行写入 data/funding_ledger.csv。31 列中英双语（表头变更自动迁移）。"""
        csv_path = Path("data/funding_ledger.csv")
        file_exists = csv_path.exists()

        # 旧→新格式迁移放在 row 构建之后进行（需按 row.keys() 比对表头），见下方写入前。

        settle_ms = self.last_funding_tx_ms or settle_fallback_ms

        row = {
            "记录时间(record_time)": self._fmt_ts(time.time()),
            "交易对(symbol)": symbol,
            "方向(direction)": direction,
            "资金费率(funding_rate)": f"{funding_rate:.6f}",
            "名义价值USDT(notional_usdt)": f"{notional:.4f}",
            "开仓价(entry_price)": f"{entry_price:.6f}",
            "发出开仓订单时间(open_send_time)": self._fmt_ts(self._open_send_time),
            "开仓成交时间(open_fill_time)": self._fmt_ms(self._open_fill_time_ms),
            "平仓价(exit_price)": f"{exit_price:.6f}",
            "滑点%(slippage_pct)": f"{slippage_pct:.4f}",
            "价格盈亏USDT(price_pnl)": f"{price_pnl:.6f}",
            "资费收入USDT(funding_income)": f"{funding_income:.6f}",
            "手续费_原生(fee_native)": f"{fee_native:.8f}",
            "手续费_币种(fee_asset)": fee_asset,
            "手续费USDT(fee_usdt)": f"{fee_usdt:.6f}",
            "总收益USDT(net_pnl)": f"{net_pnl:.6f}",
            "资费结算时间(funding_settle_time)": self._fmt_ms(settle_ms),
            "收到资费推送时间(funding_recv_time)": self._fmt_ms(self.last_funding_recv_ms),
            "发出止盈单时间(tp_send_time)": self._fmt_ts(self._tp_send_time),
            "发出平仓订单时间(close_send_time)": self._fmt_ts(self._close_send_time),
            "平仓成交时间(close_fill_time)": self._fmt_ms(self._close_fill_time_ms),
            "收到开仓确认时间(open_ack_time)": self._fmt_ts(self._open_ack_time),
            "收到平仓确认时间(close_ack_time)": self._fmt_ts(self._close_ack_time),
            "触发方式(trigger_type)": self._last_trigger_type,
            "成交笔数(fill_count)": str(fill_count),
            "资费周期h(funding_interval_hours)": str(self._interval_hours),
            "批大小(batch_size)": str(self._batch_size),
            "批内位次(subset_rank)": str(self._subset_rank),
            "结算UTC时(settle_hour_utc)": str(self._settle_hour_utc),
            "平仓死线ms(close_deadline_ms)": f"{self.current_delay_ms:.0f}",
            "是否收到资费(funding_collected)": "1" if abs(funding_income) > 1e-12 else "0",
        }

        # 旧格式 → 新格式自动迁移：表头与当前列不完全一致(缺列/改名/加列)即备份重建，永久自适应。
        # 用「复制备份」而非「改名」：先 copy 一份 bak（只读，不与 Excel 争独占锁），本次写入用 "w" 截断重写新表头。
        # xlsx 备份单独 try：即使被 Excel 占用备份失败，也不拖累 CSV 迁移（xlsx 本就每次从 CSV 整体重建）。
        migrated = False
        if file_exists:
            try:
                with open(csv_path, "r", encoding="utf-8") as f:
                    first_line = f.readline().rstrip("\r\n")
                if first_line and first_line != ",".join(row.keys()):
                    suffix = time.strftime("%Y%m%d_%H%M%S")
                    shutil.copy2(csv_path, csv_path.with_name(f"funding_ledger_bak_{suffix}.csv"))
                    try:
                        xlsx_prev = Path("data/funding_ledger.xlsx")
                        if xlsx_prev.exists():
                            shutil.copy2(xlsx_prev, xlsx_prev.with_name(f"funding_ledger_bak_{suffix}.xlsx"))
                    except Exception as ex:
                        logger.warning("[账本迁移] 旧 XLSX 备份失败(可能被 Excel 占用)，CSV 迁移继续: %s", ex)
                    migrated = True
                    file_exists = False
                    logger.info("[账本迁移] 表头变更，旧账本已备份(后缀 %s)，重建为最新格式", suffix)
            except Exception as ex:
                logger.warning("[账本迁移] 迁移失败，保持原文件不动: %s", ex)

        # 记账绝不能影响交易主流程：写入独立包裹，失败只记 error
        try:
            csv_path.parent.mkdir(parents=True, exist_ok=True)
            # 迁移轮用 "w" 截断旧内容重写新表头（旧数据已 copy 到 bak）；常规轮用 "a" 追加。
            mode = "w" if migrated else "a"
            with open(csv_path, mode, newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=list(row.keys()))
                if not file_exists:
                    writer.writeheader()
                writer.writerow(row)
        except Exception as exc:
            logger.error("[账本落盘失败] 无法写入 CSV: %s", exc)

        # 读历史，输出累计大盘（过滤空行 + 异常转型保护，防脏数据崩溃收尾）
        try:
            all_rows: list[dict] = []
            with open(csv_path, "r", encoding="utf-8") as f:
                for r in csv.DictReader(f):
                    if r.get("总收益USDT(net_pnl)"):  # 跳过断电/写入中断残留的空行
                        all_rows.append(r)
            total = len(all_rows)
            if total > 0:
                wins = sum(1 for r in all_rows if float(r.get("总收益USDT(net_pnl)", 0) or 0) > 0)
                win_rate = wins / total * 100
                total_net = sum(float(r.get("总收益USDT(net_pnl)", 0) or 0) for r in all_rows)
                ws_rows = [r for r in all_rows if r.get("触发方式(trigger_type)") == "WS_TRIGGER"]
                ws_rate = len(ws_rows) / total * 100
                logger.info("[账本大盘] 📊 累计 %d 单 | 盈利 %d 单(%.1f%%) | 累计净利 %.4fU | WS抢跑 %d 单(%.1f%%)",
                            total, wins, win_rate, total_net, len(ws_rows), ws_rate)
            # XLSX 导出：自动列宽 + 时间格式，打开即用
            self._write_ledger_xlsx(all_rows, list(row.keys()))
        except Exception as e:
            logger.debug("[账本大盘统计跳过] 解析历史数据异常(可能有脏行): %s", e)

    @staticmethod
    def _write_ledger_xlsx(all_rows: list[dict], fieldnames: list[str]) -> None:
        """从 CSV 行数据生成 XLSX：自适应列宽 + 时间列格式化为 yyyy-mm-dd hh:mm:ss.000。
        与 CSV 并行写入，失败不影响交易。"""
        if not _XLSX_AVAILABLE:
            return
        try:
            xlsx_path = Path("data/funding_ledger.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.title = "资费狙击账本"

            # 写表头（加粗 + 居中对齐）
            header_font = Font(bold=True)
            header_align = Alignment(horizontal="center", vertical="center")
            for ci, name in enumerate(fieldnames, 1):
                cell = ws.cell(row=1, column=ci, value=name)
                cell.font = header_font
                cell.alignment = header_align

            # 识别时间列（列名含"时间"或"time"）
            time_cols = {ci for ci, n in enumerate(fieldnames, 1)
                         if "时间" in n or "time" in n.lower()}

            # 写数据行
            for ri, row in enumerate(all_rows, 2):
                for ci, name in enumerate(fieldnames, 1):
                    val = row.get(name, "")
                    cell = ws.cell(row=ri, column=ci, value=val)

            # 自适应列宽：取表头长度和每列数据最长值的 max，上下限 8-40
            for ci in range(1, len(fieldnames) + 1):
                header_len = len(ws.cell(row=1, column=ci).value or "")
                data_max = 0
                for ri in range(2, len(all_rows) + 2):
                    v = ws.cell(row=ri, column=ci).value
                    if v:
                        data_max = max(data_max, len(str(v)))
                width = max(header_len * 1.3, data_max + 2)
                width = max(8, min(width, 40))
                ws.column_dimensions[get_column_letter(ci)].width = width

            wb.save(str(xlsx_path))
            logger.debug("[账本] XLSX 已更新 %d 行", len(all_rows))
        except Exception as exc:
            logger.debug("[账本] XLSX 写入失败: %s", exc)

    # ── 主循环 ─────────────────────────────────────────────

    async def run_cycle(self) -> None:
        """狙击主循环。"""
        # ── 持仓管理 ──
        if self.is_open:
            if not await self.check_position():
                return
            if self.should_exit():
                success, _ = await self.close_position_fast()
                return
            remain_ms = self.next_funding_time_ms - time.time() * 1000
            logger.info("[狙击] 持有 %s %s, 距结算 %dms", self.symbol, self.direction, int(remain_ms))
            return

        # ── 狙击模式 ──
        snipe_ms = self._next_snipe_settle_ms
        if snipe_ms:
            remain_ms = snipe_ms - time.time() * 1000
            if remain_ms <= 0:
                self._next_snipe_settle_ms = 0.0
            elif remain_ms <= CROSS_SNIPE_WINDOW_SEC * 1000:
                # 防外部定时器并发穿透：同一时刻只允许一个狙击协程
                if self.is_sniping:
                    return
                self.is_sniping = True
                try:
                    # Step 1: T-10s 扫描 + 设杠杆
                    scan_at_ms = snipe_ms - CROSS_SNIPER_SCAN_OFFSET_SEC * 1000
                    await asyncio.sleep(max(0, (scan_at_ms - time.time() * 1000) / 1000))
                    logger.info("[狙击] 狙击扫描：距结算 %dms",
                                max(0, int(snipe_ms - time.time() * 1000)))

                    bal = await self.bot._cross_get_bn_futures_balance()
                    if bal <= 0:
                        logger.error("[狙击] 余额不足 %.2f", bal)
                        return
                    position_usdt = bal * _c.BN_SNIPE_POSITION_SIZE_RATIO * CROSS_LEVERAGE
                    try:
                        df = await asyncio.wait_for(self.scan(position_usdt), timeout=12)
                    except asyncio.TimeoutError:
                        logger.warning("[狙击] 扫描超时，放弃")
                        return

                    passing = df[df["passes"] == True] if not df.empty else pd.DataFrame()
                    if passing.empty:
                        logger.info("[狙击] 无合格候选")
                        return

                    # 遍历候选直到能开仓，预计算 amount
                    chosen = None
                    chosen_amount = 0.0
                    for idx in range(min(10, len(passing))):
                        candidate = passing.iloc[idx]
                        sym = str(candidate["symbol"])
                        price = float(candidate["futures_price"])
                        amount = await self.bot._cross_calculate_amount("binance", sym, price, position_usdt)
                        if amount > 0:
                            chosen = candidate
                            chosen_amount = amount
                            logger.info("[狙击] 目标 #%d %s: 费率=%.4f%% 方向=%s 张数=%.4f",
                                         idx + 1, sym, float(candidate["funding_rate"]) * 100,
                                         str(candidate["direction"]), chosen_amount)
                            break
                        logger.warning("[狙击] #%d %s 数量=%.4f，尝试下一位", idx + 1, sym, amount)

                    if chosen is None:
                        logger.info("[狙击] 所有候选数量不足")
                        return

                    # T-10s 预处理: 设杠杆 + 动态延时矩阵
                    await self.bot._binance_set_leverage(str(chosen["symbol"]))

                    rate_abs = abs(float(chosen["funding_rate"]))
                    nft_ms = float(chosen["next_funding_time_ms"])
                    settle_hour = time.localtime(nft_ms / 1000).tm_hour
                    interval_hours = int(chosen.get("interval_hours", DEFAULT_FUNDING_INTERVAL_HOURS))
                    self._interval_hours = interval_hours
                    self._batch_size = int(chosen.get("batch_size", 0))
                    self._subset_rank = int(chosen.get("subset_rank", -1))
                    self._settle_hour_utc = time.gmtime(nft_ms / 1000).tm_hour

                    # ── 按字母排位线性设开仓/平仓时点（默认）或固定值 ──
                    # rank% 越大(字母越靠后)→记账越晚→开仓延迟与平仓死线都线性增大。
                    # TradFi 股票合约例外：T实测~107ms，早于普通币，需更早开仓。
                    contract_type = str(chosen.get("contract_type", ""))
                    is_tradfi = contract_type.startswith("TRAD")  # 币安实际返回 TRADIFI_PERPETUAL（官方拼写多个I），前缀匹配兜住两种拼法
                    rank_pct = float(chosen.get("rank_pct", 0.0))
                    p = min(1.0, max(0.0, rank_pct))
                    if _c.BN_SNIPE_RANK_TIMING_ENABLED:
                        if is_tradfi:
                            open_offset_ms = -_c.BN_SNIPE_TRADFI_OPEN_MS
                        else:
                            open_offset_ms = -(_c.BN_SNIPE_OPEN_MS_MIN
                                               + p * (_c.BN_SNIPE_OPEN_MS_MAX - _c.BN_SNIPE_OPEN_MS_MIN))
                        self.current_delay_ms = (_c.BN_SNIPE_CLOSE_MS_MIN
                                                 + p * (_c.BN_SNIPE_CLOSE_MS_MAX - _c.BN_SNIPE_CLOSE_MS_MIN))
                    else:
                        open_offset_ms = _c.BN_SNIPE_OPEN_OFFSET_MS
                        self.current_delay_ms = 400.0

                    logger.info("[狙击] 目标 %s | 节点 %02d:00 | 周期 %dh | 妖币=%s | 字母排位=%.0f%% | 开=结算后%.0fms 平死线=%.0fms",
                                 str(chosen["symbol"]), settle_hour, interval_hours,
                                 "YES" if rate_abs >= 0.015 else "NO",
                                 rank_pct * 100, abs(open_offset_ms), self.current_delay_ms)

                    fast_args = {
                        "symbol": str(chosen["symbol"]),
                        "direction": str(chosen["direction"]),
                        "amount": chosen_amount,
                        "price": float(chosen["futures_price"]),
                        "nft": nft_ms,
                        "rate": float(chosen["funding_rate"]),
                        "net_rate": float(chosen["net_rate"]),
                        "futures_fee": float(chosen["futures_taker_fee"]),
                    }

                    # Step 2: T-2s 战术网络热身 → T-1s 开仓
                    warmup_at_ms = snipe_ms - 2000
                    remain_to_warmup = warmup_at_ms - time.time() * 1000
                    if remain_to_warmup > 0:
                        await asyncio.sleep(remain_to_warmup / 1000)
                        try:
                            await asyncio.wait_for(self.bot.futures.fetch_time(), timeout=2)
                        except Exception:
                            pass

                    open_at_ms = snipe_ms - open_offset_ms
                    tag = "结算前%dms" % open_offset_ms if open_offset_ms > 0 else ("结算后%dms" % abs(open_offset_ms) if open_offset_ms < 0 else "整点")
                    logger.info("[狙击] 锁定 %s，开仓=%s（+%dms）",
                                 fast_args["symbol"], tag, int(open_at_ms - snipe_ms))
                    await asyncio.sleep(max(0, (open_at_ms - time.time() * 1000) / 1000))

                    # 终极防线: 关闭 GC, 开仓→平仓这 1s 内绝不允许 GC 背刺
                    gc.disable()
                    try:
                        # 开仓前擦净信号画布：确保 open 内或 open 后任何时刻到达的
                        # FUNDING_FEE 都能被 wait_for 捕获，避免 clear() 放在 open
                        # 之后抹杀已到达的 WS 信号。
                        self.ws_funding_arrived_event.clear()
                        # 诊断: 结算±5s窗口内转储全部WS原始JSON，定位FUNDING_FEE根因
                        self.bot._funding_raw_dump_until = time.time() + 10
                        ok = await self.open_position_fast(fast_args)

                        if ok:
                            # Step 3: WS资费到账 + 硬死线超时 双轨竞赛平仓
                            hard_deadline = (snipe_ms + self.current_delay_ms) / 1000.0
                            timeout = max(0.001, hard_deadline - time.time())

                            # 断线预警: 资费WS若在开仓后断开，只会走硬超时兜底
                            lost_at = getattr(self.bot, "_funding_ws_lost_at", 0.0)
                            if lost_at and time.time() - lost_at < 30:
                                logger.warning("[狙击] 资费WS断线中 (%.0fs前)，将依赖硬死线兜底",
                                               time.time() - lost_at)

                            try:
                                await asyncio.wait_for(
                                    self.ws_funding_arrived_event.wait(), timeout=timeout,
                                )
                                if self.tp_filled:
                                    # 事件被止盈成交提前拉响（非资费到账）→ 诚实标注，免误报捷报
                                    logger.info("[狙击] 止盈已在机房侧成交，提前唤醒收尾 | 跳过资费等待")
                                    self._last_trigger_type = "TP_止盈"
                                    self._last_e2e_ms = 0
                                else:
                                    fire_wall_ms = int(time.time() * 1000)
                                    tx_ms = getattr(self, "last_funding_tx_ms", 0)
                                    total_e2e = fire_wall_ms - tx_ms if tx_ms else 0
                                    logger.info("[狙击] WS资费信号捷报！提前 %.0fms 触发平仓 | 穿透全链路(币安记账→平仓发单)总耗时: %dms",
                                                 (hard_deadline - time.time()) * 1000, total_e2e)
                                    self._last_trigger_type = "WS_TRIGGER"
                                    self._last_e2e_ms = total_e2e
                            except asyncio.TimeoutError:
                                logger.warning("[狙击] WS超时，硬时钟强制执行平仓 | 死线已到")
                                self._last_trigger_type = "TIMEOUT_兜底"
                                self._last_e2e_ms = 0

                            closed, pure_rtt = await self.close_position_fast()
                            logger.info("[延迟] 平仓纯净 RTT: %.1fms", pure_rtt)
                    finally:
                        gc.enable()
                finally:
                    self.is_sniping = False
                    self._next_snipe_settle_ms = 0.0
                return

        # ── 常规扫描 ──
        bal = await self.bot._cross_get_bn_futures_balance()
        self._dash_balance = bal
        if bal <= 0:
            logger.error("[狙击] 余额不足 %.2f，等待", bal)
            return
        position_usdt = bal * _c.BN_SNIPE_POSITION_SIZE_RATIO * CROSS_LEVERAGE
        logger.info("┏━━ [狙击] 主扫开始 | 余额=%.2f | 可用仓位=%.2f ━━", bal, position_usdt)
        df = await self.scan(position_usdt)
        self._dash_df = df

        if df.empty:
            logger.info("[狙击] 无符合条件的币种。")
            return

        passing = df[df["passes"] == True] if "passes" in df.columns else pd.DataFrame()
        self._print_opportunity_table(df, position_usdt)

        if not passing.empty:
            # scan 已按 [结算时间↑, 费率↓] 排好序，直接取首位
            best = passing.iloc[0]

            self._next_snipe_settle_ms = float(best["next_funding_time_ms"])
            settle_local = str(best.get("settle_local", "?"))
            remain_min = float(best.get("remain_min", 0) or 0)
            remain_str = f"{int(remain_min//60)}h{int(remain_min%60):02d}m" if remain_min >= 60 else f"{int(remain_min)}m"
            logger.info("[狙击] 下次狙击目标: %s 费率=%.4f%% 方向=%s 结算=%s(北京时间) 还剩%s",
                         str(best["symbol"]), float(best["funding_rate"]) * 100,
                         str(best["direction"]), settle_local, remain_str)
        else:
            self._next_snipe_settle_ms = 0.0

        # 若刚发现机会且结算在 2*snipe_window 内 → 直接跳狙击
        if not self._snipe_loop_guard and self._next_snipe_settle_ms:
            remain_ms = self._next_snipe_settle_ms - time.time() * 1000
            if 0 < remain_ms <= CROSS_SNIPE_WINDOW_SEC * 1000 * 2:
                logger.info("[狙击] 距结算仅 %dms，直接进入狙击流程", int(remain_ms))
                self._snipe_loop_guard = True
                try:
                    await self.run_cycle()
                finally:
                    self._snipe_loop_guard = False

    def _print_opportunity_table(self, df: pd.DataFrame, position_usdt: float) -> None:
        """打印两张机会表格：时间优先 TOP3 + 费率优先 TOP3。"""
        if df.empty:
            return

        t1 = df.sort_values(["next_funding_time_ms", "abs_rate"], ascending=[True, False]).head(3)
        t2 = df.sort_values("abs_rate", ascending=False).head(3)

        logger.info("  ═══════════════════════════════════════════════════════════════════")
        logger.info("  币安资金费率狙击 (裸合约, 无对冲) — 可用 ≈%.0f USDT", position_usdt)
        logger.info("  ═══════════════════════════════════════════════════════════════════")

        self._print_table("按结算时间优先 (TOP 3)", t1)
        self._print_table("按费率绝对值优先 (TOP 3)", t2)

    def _print_table(self, title: str, top: pd.DataFrame) -> None:
        """打印单张机会表。"""
        if top.empty:
            return

        def _fmt_vol(v: float) -> str:
            if v >= 1_000_000:
                return f"{v/1e6:.1f}M"
            elif v >= 1_000:
                return f"{v/1e3:.0f}K"
            return f"{v:.0f}"

        def _fmt_remain(m: float) -> str:
            if m >= 60:
                return f"{int(m//60)}h{int(m%60):02d}m"
            return f"{int(m)}m"

        logger.info("  ── %s ──", title)
        logger.info("    %-10s %9s %8s %7s %7s %5s %7s %7s %6s %5s",
                    "币种", "价格", "费率%", "abs%", "净收益%", "方向",
                    "24h成交", "距结算", "结算", "通过")
        logger.info("  ------------------------------------------------------------------")
        for _, row in top.iterrows():
            price_f = float(row["futures_price"])
            vol = float(row.get("quote_volume", 0) or 0)
            remain_m = float(row.get("remain_min", 0) or 0)
            settle = str(row.get("settle_local", "?"))
            logger.info("    %s%-10s %9s %7.2f%% %6.2f%% %6.2f%% %5s %7s %7s %6s %5s",
                        "✓" if row["passes"] else " ",
                        str(row["base"]),
                        f"{price_f:.6f}" if price_f < 1 else f"{price_f:.4f}",
                        float(row["funding_rate"]) * 100,
                        float(row["abs_rate"]) * 100,
                        float(row["net_rate"]) * 100,
                        str(row["direction"]),
                        _fmt_vol(vol),
                        _fmt_remain(remain_m),
                        settle,
                        "✓" if row["passes"] else "✗")
        logger.info("")
